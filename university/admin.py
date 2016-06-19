# -*- coding: utf-8 -*-
import sys

from django.contrib import admin
from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from .models import University, Faculty, Indicator, FacultyIndicators, IndicatorIntervals




class FacultyAdmin(admin.ModelAdmin):
    fields = ['name', 'shortcut', 'university', 'faculty_managers']
    list_display = ('name', 'shortcut', 'total_rank', 'university')

    def get_queryset(self, request):
        qs = super(FacultyAdmin, self).get_queryset(request)
        if request.user.is_superuser:
                return qs
        user_universities = request.user.university_set.all().values_list('id', flat=True)
        print(user_universities)
        uni_fac_set = Faculty.objects.exclude(university__id__in=user_universities).values_list('id', flat=True)  #user_university[1]
        print(uni_fac_set)
        return qs.exclude(pk__in = uni_fac_set)


class IndicatorAdmin(admin.ModelAdmin):
    list_display = ('shortcut', 'is_public', 'pub_date')

class IntervalAdmin(admin.ModelAdmin):
    list_display = ('get_name', 'comment')
    def get_name(self, obj):
        return obj.__str__()
    get_name.short_description = 'Okres'
  #  get_author.admin_order_field = 'book__author'

class RelatedFacultyFilter(admin.SimpleListFilter):
    # Human-readable title which will be displayed in the
    # right admin sidebar just above the filter options.
    title = _('Wydział')

    parameter_name = 'faculty'

    def lookups(self, request, model_admin):
        if request.user.is_superuser:
            fall=Faculty.objects.all()
            return [(f.id, f.name) for f in fall]
        l = []
        for g in request.user.faculty_set.all():
            l.append((g.id, g.name))
        return l

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(faculty_id=self.value())
        return queryset

class RelatedUniversityFilter(admin.SimpleListFilter):
    # Human-readable title which will be displayed in the
    # right admin sidebar just above the filter options.
    title = _('uniwersytet')

    parameter_name = 'university'

    def lookups(self, request, model_admin):
        if request.user.is_superuser:
            uall=University.objects.all()
            return [(u.id, u.name) for u in uall]
        l = []
        for g in request.user.university_set.all():
            l.append((g.id, g.name))
        return l

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(id=self.value())
        return queryset


class UniversityAdmin(admin.ModelAdmin):
    list_display = ('name', 'total_rank', 'address')
    list_filter = [RelatedUniversityFilter]

class CorrespondingDateFilter(admin.SimpleListFilter):
    # Human-readable title which will be displayed in the
    # right admin sidebar just above the filter options.
    title = _('Okres')

    parameter_name = 'time_interval'

    def lookups(self, request, model_admin):
        l=[]
        dates = IndicatorIntervals.objects.all()
        for d in dates:
            l.append((d.id, d))
        return l

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(time_interval_id=self.value())


class FacultyIndicatorsAdmin(admin.ModelAdmin):
    def get_readonly_fields(self, request, obj=None):
        if obj and not request.user.is_superuser:
            return ['faculty', 'indicator'] # Return a list or tuple of readonly fields' names
        return []

#    readonly_fields= []
    fields = ['indicator', 'faculty', 'value', 'significance_coefficient', 'ease_coefficient', 'time_interval', 'pub_date']
    list_display = ('indicator', 'faculty', 'value', 'time_interval')
    list_filter = (RelatedFacultyFilter, 'value', 'indicator', CorrespondingDateFilter)
    search_fields = ['indicator__name']
    list_per_page = 20



    def get_queryset(self, request):
        qs = super(FacultyIndicatorsAdmin, self).get_queryset(request)
        if request.user.is_superuser:
                return qs
        user_university = request.user.university_set.all()
        uni_fac_set = Faculty.objects.filter(id__in=user_university)  #user_university[1]
        #show faculties only from our university
        fl = FacultyIndicators.objects.exclude(faculty__id__in=request.user.faculty_set.all()).values_list('id', flat=True)
        ul = FacultyIndicators.objects.filter(faculty__id__in=uni_fac_set)
        s = set()
        s |= set(fl)
        s |= set(ul)
        print(s)
        return qs.exclude(pk__in=fl)

    def export_csv(modeladmin, request, queryset):
        import csv
        from django.utils.encoding import smart_str
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=facultyIndicators.csv'
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow([
            smart_str(u"ID"),
            smart_str(u"faculty"),
            smart_str(u"Indicator"),
            smart_str(u"Value"),
            smart_str(u"Start date"),
            smart_str(u"Finish date"),
        ])
        for obj in queryset:
            writer.writerow([
                smart_str(obj.pk),
                smart_str(obj.faculty),
                smart_str(obj.indicator),
                smart_str(obj.value),
                smart_str(obj.time_interval.start_date),
                smart_str(obj.time_interval.end_date),
            ])
        return response

    def export_xlsx(modeladmin, request, queryset):
        import openpyxl
        from openpyxl.cell import get_column_letter
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=facultyIndicators.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "MyModel"

        row_num = 0

        columns = [
            (u"ID", 15),
            (u"Faculty", 40),
            (u"Indicator", 40),
            (u"Value", 10),
            (u"Start date", 15),
            (u"End date", 15),
            (u"Comment", 15),
        ]

        for col_num in xrange(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for obj in queryset:
            row_num += 1
            row = [
                obj.pk,
                str(obj.faculty),
                str(obj.indicator),
                obj.value,
                obj.time_interval.start_date,
                obj.time_interval.end_date,
                obj.comment,
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.style.alignment.wrap_text = True

        wb.save(response)
        return response

    actions = [export_csv, export_xlsx]
    export_xlsx.short_description = u"Eksport do XLSX"
    export_csv.short_description = u"Eksport do CSV"


admin.site.register(University, UniversityAdmin)
admin.site.register(Faculty, FacultyAdmin)
admin.site.register(Indicator, IndicatorAdmin)
admin.site.register(FacultyIndicators, FacultyIndicatorsAdmin)
admin.site.register(IndicatorIntervals, IntervalAdmin)

